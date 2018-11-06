import openpyxl

#make a class to find specfic data
class Store:
    def __init__(self, store, presales, gosales):
        self.store = store
        self.presales = presales
        self.gosales = gosales

#funtion to check lists are compatible by first checking the length on the list are the same. 
def check(sales_list, media_list):
    length_sales_list = len(sales_list), length_media_list = len(media_list)
    media_false = list(), media_true = list() 
    while length_sales_list == length_media_list: #because the way i have iterated it they should be the same length
        for i in range(length_sales_list): #could of been b didn't matter
            store_name_sales = sales_list[i][0].lower(), store_name_media = media_list[i][0].lower() 
            while store_name_sales == store_name_media:
                length_sales_store_list = len(sales_list[i])
                for j in range(1, length_sales_store_list):
                    for k in range(2):
                        weeks = media_list[i][j]
                        if not str(weeks[k]).isnum():
                            print(f"file media has an error on row {i}, on the {} column")
                            break
                        elif weeks[1] == 0:
                            temp = sales_list[i][weeks[0] - 12: weeks[0] + 3]
                            media_False.append(temp)
                        elif weeks[0] > weeks[1]:
                                print(f"check correct weeks have been input on row {i} for media start and finish")
                        else:
                            temp = sales_list[i][weeks[0] - 12: weeks[0] + (weeks[1] - weeks[0])]
                            media_true.append(temp)
                        media_true.append(media_false)
                return media_true #a list that sperates my data so the 0 index is media sales, then in that list it contains string for store, 
            #and a list of at least 15 lists containing 2 csv, 1 for week, 1 for sales 

#function to create a list within a list to iterate though the week values easier
def liststore(sheets, rnum, calpha):
    v1 = list(), ListByStoreName = list(), v3 = list()
    for RowRead in range(2, rnum + 1): # becuase title doesn't matter and in excel it starts from 1 not 0
        ListByStoreName.append(sheets.cell(RowRead, 1))
        for ColumnRead in range(2, calpha + 1): #because we are looking at asales and weeks only
            value = sheet.cell(RowRead, ColumnRead)
            while value == NULL:
                value = 0
            while len(v1) == 2:
                ListByStoreName.append(v1)
                v1 = list()
            if sheets.cell(RowRead, 1) == sheets.cell(RowRead - 1, 1):
                v1.append(value)
            else:
                v3.append(ListByStoreName)
                ListByStoreName = list()
    return v3

#function to check if file titles are correct
def correct(cellname, version):
    name = list() 
    for icn in cellname:
        name += icn.title()    
    if name == version:
        return True
    else:
        return False

#function to automate what are in the title cells for each spreadsheet    
def titlemaker(sheet, col):
    cellnames = list()
    for inc in range(1, col + 1):
        temp = sheet.cell(1, inc)
        if not temp.isalpha():
            print("invalid entry on file")
            break
        else:
            cellnames += temp
    return cellnames

#xyz & xyz1 = readable[mediaside] or readable[nonside]
def readSale(SaleCSV):
    lengthCSV = len(SaleCSV)
    for irsale in range(lengthCSV):
        LocationSalesLength = len(SaleCSV[irsale])
        for jrsale in range(1, LocationSaleLength):
            sale = SaleCsv[irsale][jrsale][1::2] # [1::2] means only get the values on the right.
            if sale.isnum():
                sale = int(sale)
            return sale
            
def readYAXIS(WeekRead):
    NumStores = len(WeekRead)
    for irweek in range(NumStores):
        NumWeeks = len(WeekRead[irweek])
        for jrweek in range(1, Numweeks):
            ExtractedWeeks = WeekRead[irweek][jrweek][::2] # means values on left)
            return ExtractedWeeks
        
 def readXAxis(StoreNameLoc):
    NumberStores = len(StoreNameLoc)
    for iraxis in range(NuymberStores):
        ExtractedStores = StoreNameLoc[iraxis][0]
        return ExtractedStores  

# this function checks the axis entry if it is numeric we know that it belongs to the x-axis/columns or else it belongs in the y-axis
# rows. which means i would literally go down the rows if it was the y-axis, and accross the columns if the x-axis 
def writeAxis(ReadAxis, sheetx):
    lowerxy = str(ReadAxis).lower()
    for axisi in range(1, len(ReadAxis) + 1):
    if lowerxy.isnum():
            Axis = "Week ", sheetx.cell(axisi + 1, 1)
    else:
            Axis = sheetx.cell(1, axisi + 1)
    return Axis
    
# so im going to call the value of either media(readable[0]) or non media(readable[1]) in the function readSale with the argument of 
# the sheet im working on them im going to find the length of the lists in the data set(since it's excel we need 2 add 1 to each value
# in the range, the first one as to not include the store name. then after the string name. each index should contain a list of size 2
# therefore i must be consider the sytle of the data, which i intended to plot the prelaunch week
def writeSale(SaleRead, WriteSheet):
    lengthStoreInSalesRead = len(SaleRead)
    for WriteRow in range(1, LengthStoresInSalesRead + 1):
        NumWeeksRead = len(SalesRead[WriteRow])
        for WriteColumn in range(1, NumWeeksRead + 1):
            return WriteSheet.cell(WriteRow, WriteColumn).value(SalesRead)
        
def whichsheet(argx, as1, as2):
    if argx == mediaside:
        Wsheet = asheet
    elif argx == nonmediaside:
        Wsheet = asheet1
    else:
        print(f"unknownn variable {argx} entered")
        break
    return Wsheet

writeSale(readSale(argx, Wsheet)), writeAxis(readYAxis(argx, Wsheet)), writeAxis(readYaxis(argx, Wsheet)))        
#creating a variable to check if true or false
mc = correct(titlemaker(mhm, mcolumns), campaignform), sc = correct(titlemaker(shm, scolumns), salesform)

#creating variables that contain string to the desired locations
sales = "C:\\Sales.xlsx", campaign = "C:\\Campaign.xlsx"

# load workbook, files must be called 'Sales.xlsx' and 'campaign.xlsx' on the c drive
sa = openpyxl.load_workbook(sales), md = openpyxl.load_workbook(campaign)

# creating object of the sheet i want to work on
shs = sa[0], shm = md[0]

#find rows & coiumns
srows = shs.max_row, scolumns = shs.max_column
mrows = shm.max_row, mcolumns = shm.max_column

campaignform = ["Store", "Start", "Finish"]
salesform = ["Store", "Week", "Sales"]

#making the data readable for me the user
if mc False:
    print(f"ERROR: file {campaign} doesn't appear to be formatted correctly")
    break
elif sc False:
    print(f"ERROR: file {sales} doesn't appear to be formatted correctly")
    break
else:
    readable = check(liststore(shs, srows, scolumns), liststore(shm, mrows, mcolumns))
    mediaside = readable[0], nonmediaside = readable[1]
    

#preparing my sheets and naming them
wk = openpyxl.Workbook()
asheet = wk.active
asheet.title = "Media Performance"

#naming extra sheets in range of 3 because i am creating 3 new sheets.
for sheetnum in range (3):
    if sheetnum == 0:
        sheetTitle = "Analysis"
        sheet0 = sheetTitle
    elif sheetnum == 1:
        sheetTitle = "Non-Media Performance"
        sheet1 = sheetTitle
    else:
        sheetTitle = "Visulisation"
        sheet2 = sheetTitle
    wk.create_sheet = (title = sheetTitle)


                                                                                             
brandname = input("please enter the name of the product you intend to be using here : ")
#saving file
wk.save(f"C:\\{brandname}anaylsis.xlsx")
